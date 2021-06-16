# coding: utf8
from docx import Document
from docx.oxml.ns import qn
from docx.shared import Pt, RGBColor
from openpyxl.styles import Border, Side, Font
import pymysql
import time
import openpyxl
from openpyxl import Workbook
import shutil
import os
from ftplib import FTP


def ftp_upload(cwd, src, dst):
    try:
        ftp = FTP()
        ftp.connect("13.32.4.176", 21)  # 第一个参数可以是ftp服务器的ip或者域名，第二个参数为ftp服务器的连接端口，默认为21
        ftp.login('files', '')  # 匿名登录直接使用ftp.login()
        # ftp.cwd('县域视频数据统计通报'.encode('gbk').decode('iso-8859-1'))
        ftp.cwd(cwd.encode('gbk').decode('iso-8859-1'))
        fp = open(src, 'rb')
        ftp.storbinary('STOR ' + dst.encode('gbk').decode('iso-8859-1'), fp, 1024)
        print(str(dst) + '上传ftp成功')
    except Exception as e:
        print('上传ftp异常：' + str(e))
    finally:
        ftp.close()  # 关闭ftp


mydb = pymysql.connect(
    host='13.32.4.170',
    # host='192.168.23.112',
    port=3306,
    user='root',
    password='wanfang@2001',
    db='db1400',
    # db='ligh',
    charset='utf8'
)


def mysql_select(mydb, sql):
    res_select = ()
    try:
        # print('查询...')
        cursor = mydb.cursor()
        cursor.execute(sql)
        res_select = cursor.fetchall()
    except Exception as e:
        print("查询异常:" + str(e))
    finally:
        cursor.close()
        return res_select


def open_xlsx(file):
    workbook = None
    try:
        # 打开excle文件，获取工作簿对象
        workbook = openpyxl.load_workbook(file)
    except Exception as e:
        print(file + '打开异常：' + str(e))
    finally:
        return workbook


def copy_file(src_file, dst_file):
    if not os.path.exists(dst_file):
        shutil.copy(src_file, dst_file)
        print(str(dst_file) + "创建成功！")


if __name__ == '__main__':
    border = Border(left=Side(border_style='thin', color='000000'), right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'), bottom=Side(border_style='thin', color='000000'))
    ftp_cwd = '智慧社区无效统计'
    today = time.strftime("%Y-%m-%d", time.localtime())
    today_dir = os.path.join('C:\\Users\\wf\\Desktop', str(today))
    # today_dir = os.path.join('/home/ligh/relay_stat', str(today))
    area_list = ['鹿泉区', '井陉县', '藁城区', '平山县', '行唐县', '无极县', '高邑县', '栾城区', '矿区', '长安区', '循环化工园区', '晋州市', '赵县', '高新区', '裕华区', '新华区', '深泽县', '正定县', '灵寿县', '新乐市', '元氏县', '桥西区', '赞皇县']
    if not os.path.exists(today_dir):
        os.mkdir(today_dir)

    for area in area_list:
        workbook_zhihui = Workbook()  # 创建工作簿

        # '''
        # '汇总'sheet页初始化
        sheet_collect = workbook_zhihui.create_sheet("汇总")
        sheet_collect['A1'] = '序号'
        sheet_collect.merge_cells('A1:A2')
        sheet_collect['B1'] = '小区名称'
        sheet_collect.merge_cells('B1:B2')
        sheet_collect['C1'] = '人脸'
        sheet_collect.merge_cells('C1:E1')
        sheet_collect['F1'] = '车辆'
        sheet_collect.merge_cells('F1:H1')
        sheet_collect['C2'] = '总数'
        sheet_collect['D2'] = '有效'
        sheet_collect['E2'] = '无效'
        sheet_collect['F2'] = '总数'
        sheet_collect['G2'] = '有效'
        sheet_collect['H2'] = '无效'

        sheet_collect['A1'].border = border
        sheet_collect['B1'].border = border
        sheet_collect['C1'].border = border
        sheet_collect['F1'].border = border
        sheet_collect['C2'].border = border
        sheet_collect['D2'].border = border
        sheet_collect['E2'].border = border
        sheet_collect['F2'].border = border
        sheet_collect['G2'].border = border
        sheet_collect['H2'].border = border

        # '汇总'sheet页写入
        sql_collect = "select xqmc,sum(a) a,sum(b) b,sum(a) - sum(b) cz1,sum(c) c,sum(d) d,sum(c) - sum(d) cz2 from (select a.xqmc, a.bm, count(a.ape_id) a, count(b.ape_id) b, 0 c, 0 d from (select b.xqmc, b.bm, a.ape_id from ape a join dm_xq b on a.place_code = b.bm where b.xian = '" + area + "' and a.function_type = '2') a left join (select a.device_id ape_id from t_push_data_log a where a.date = curdate() and a.type = '12' and a.result = '1' group by a.device_id) b on a.ape_id = b.ape_id group by a.xqmc, a.bm union all select a.xqmc, a.bm, 0 a, 0 b, count(a.tollgate_id) c, count(b.tollgate_id) d from (select b.xqmc, b.bm, a.tollgate_id from tollgate a join dm_xq b on a.place_code = b.bm where b.xian = '" + area + "') a left join (select a.tollgate_id from t_push_data_log a where a.date = curdate() and a.type = '13' and a.result = '1' group by a.tollgate_id) b on a.tollgate_id = b.tollgate_id group by a.xqmc, a.bm) a group by a.a.xqmc, a.bm order by a.bm"
        res_collect = mysql_select(mydb, sql_collect)
        if res_collect:
            collect_num = 3
            for line in res_collect:
                sheet_collect['A' + str(collect_num)] = int(collect_num - 2)
                sheet_collect['B' + str(collect_num)] = str(line[0])
                sheet_collect['C' + str(collect_num)] = int(line[1])
                sheet_collect['D' + str(collect_num)] = int(line[2])
                sheet_collect['E' + str(collect_num)] = int(line[3])
                sheet_collect['F' + str(collect_num)] = int(line[4])
                sheet_collect['G' + str(collect_num)] = int(line[5])
                sheet_collect['H' + str(collect_num)] = int(line[6])

                sheet_collect['A' + str(collect_num)].border = border
                sheet_collect['B' + str(collect_num)].border = border
                sheet_collect['C' + str(collect_num)].border = border
                sheet_collect['D' + str(collect_num)].border = border
                sheet_collect['E' + str(collect_num)].border = border
                sheet_collect['F' + str(collect_num)].border = border
                sheet_collect['G' + str(collect_num)].border = border
                sheet_collect['H' + str(collect_num)].border = border
                collect_num += 1
                
        # '''
        # '无效设备'、'无效卡口'sheet页初始化
        sheet_detail_ape = workbook_zhihui.create_sheet("无效设备")
        sheet_detail_tollgate = workbook_zhihui.create_sheet("无效卡口")
        sheet_detail_ape['A1'] = '序号'
        sheet_detail_ape['B1'] = '小区名称'
        sheet_detail_ape['C1'] = '设备编码'
        sheet_detail_ape['D1'] = '设备名称'
        sheet_detail_ape['A1'].border = border
        sheet_detail_ape['B1'].border = border
        sheet_detail_ape['C1'].border = border
        sheet_detail_ape['D1'].border = border

        sheet_detail_tollgate['A1'] = '序号'
        sheet_detail_tollgate['B1'] = '小区名称'
        sheet_detail_tollgate['C1'] = '设备编码'
        sheet_detail_tollgate['D1'] = '设备名称'
        sheet_detail_tollgate['A1'].border = border
        sheet_detail_tollgate['B1'].border = border
        sheet_detail_tollgate['C1'].border = border
        sheet_detail_tollgate['D1'].border = border

        # '无效设备'、'无效卡口'sheet页写入
        sql_detail = "select b.xqmc, a.ape_id, a.name, '人脸' a from ape a join dm_xq b on a.place_code = b.bm and b.xian = '" + area + "' where a.function_type = '2' and not exists (select 1 from (select a.device_id ape_id from t_push_data_log a where a.date = curdate() and a.type = '12' and a.result = '1' group by a.device_id) b where a.ape_id = b.ape_id) union ALL select b.xqmc, a.tollgate_id, a.name, '卡口' a from tollgate a join dm_xq b on a.place_code = b.bm and b.xian = '" + area + "' where not exists (select 1 from (select a.tollgate_id from t_push_data_log a where a.date = curdate() and a.type = '13' and a.result = '1' group by a.tollgate_id) b where a.tollgate_id = b.tollgate_id)"
        res_detail = mysql_select(mydb, sql_detail)
        if res_detail:
            ape_num = 2
            tollgate_num = 2
            for line in res_detail:
                if line[3] == '人脸':
                    sheet_detail_ape['A' + str(ape_num)] = int(ape_num - 1)
                    sheet_detail_ape['B' + str(ape_num)] = str(line[0])
                    sheet_detail_ape['C' + str(ape_num)] = str(line[1])
                    sheet_detail_ape['D' + str(ape_num)] = str(line[2])

                    sheet_detail_ape['A' + str(ape_num)].border = border
                    sheet_detail_ape['B' + str(ape_num)].border = border
                    sheet_detail_ape['C' + str(ape_num)].border = border
                    sheet_detail_ape['D' + str(ape_num)].border = border
                    ape_num += 1

                elif line[3] == '卡口':
                    sheet_detail_tollgate['A' + str(tollgate_num)] = int(tollgate_num - 1)
                    sheet_detail_tollgate['B' + str(tollgate_num)] = str(line[0])
                    sheet_detail_tollgate['C' + str(tollgate_num)] = str(line[1])
                    sheet_detail_tollgate['D' + str(tollgate_num)] = str(line[2])

                    sheet_detail_tollgate['A' + str(tollgate_num)].border = border
                    sheet_detail_tollgate['B' + str(tollgate_num)].border = border
                    sheet_detail_tollgate['C' + str(tollgate_num)].border = border
                    sheet_detail_tollgate['D' + str(tollgate_num)].border = border
                    tollgate_num += 1

        # 保存、关闭excel文件
        workbook_zhihui.remove(workbook_zhihui['Sheet'])
        workbook_zhihui.save(os.path.join(today_dir, area + '.xlsx'))
        workbook_zhihui.close()
        print(area + 'ok')

    '''
    #‘23区县汇总’excel表
    workbook_all = Workbook()  # 创建工作簿
    # '汇总'sheet页初始化
    sheet_all = workbook_all.create_sheet("23区县汇总")
    sheet_all['A1'] = '序号'
    sheet_all['B1'] = '区县'
    sheet_all['C1'] = '无数据小区（个）'
    sheet_all['D1'] = '无数据人脸（路）'
    sheet_all['E1'] = '无数据卡口（路）'

    sheet_all['A1'].border = border
    sheet_all['B1'].border = border
    sheet_all['C1'].border = border
    sheet_all['D1'].border = border
    sheet_all['E1'].border = border

    sql_all = "select xian, count(distinct a.bm) xq, sum(a) a, sum(b) b from (select b.xian, b.bm, 1 a, 0 b from ape a join dm_xq b on a.place_code = b.bm where a.function_type = '2' and not exists (select 1 from (select a.device_id ape_id from t_push_data_log a where a.date = date(now()) and a.type = '12' and a.result = '1' group by a.device_id) b where a.ape_id = b.ape_id) union ALL select b.xian, b.bm, 0 a, 1 b from tollgate a join dm_xq b on a.place_code = b.bm where not exists (select 1 from (select a.tollgate_id from t_push_data_log a where a.date = date(now()) and a.type = '13' and a.result = '1' group by a.tollgate_id) b where a.tollgate_id = b.tollgate_id)) a group by a.xian"
    res_all = mysql_select(mydb, sql_all)
    if res_all:
        all_num = 2
        for line in res_all:
            sheet_all['A' + str(all_num)] = int(all_num - 1)
            sheet_all['B' + str(all_num)] = str(line[0])
            sheet_all['C' + str(all_num)] = int(line[1])
            sheet_all['D' + str(all_num)] = int(line[2])
            sheet_all['E' + str(all_num)] = int(line[2])

            sheet_all['A' + str(all_num)].border = border
            sheet_all['B' + str(all_num)].border = border
            sheet_all['C' + str(all_num)].border = border
            sheet_all['D' + str(all_num)].border = border
            sheet_all['E' + str(all_num)].border = border
            all_num += 1
    # 保存、关闭excel文件
    workbook_all.remove(workbook_all['Sheet'])
    workbook_all.save(os.path.join(today_dir, '23区县汇总.xlsx'))
    workbook_all.close()
    print('23区县汇总ok')
    '''


# '市区视频数据统计通报'.encode('gbk').decode('iso-8859-1')
# excel上传ftp
# todo
ftp_upload(ftp_cwd, today_dir, today_dir)
mydb.close()  # 关闭mysql连接
print('统计完成!')
