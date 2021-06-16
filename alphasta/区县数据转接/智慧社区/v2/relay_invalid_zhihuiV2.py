# coding: utf8
from openpyxl.styles import Border, Side, Font
import pymysql
import time
import openpyxl
from openpyxl import Workbook
import shutil
import os
from ftplib import FTP
import datetime
import json
import zipfile


def ftp_upload(cwd, src, dst):
    try:
        ftp = FTP()
        ftp.connect("13.32.4.176", 21)  # 第一个参数可以是ftp服务器的ip或者域名，第二个参数为ftp服务器的连接端口，默认为21
        ftp.login('files', '')  # 匿名登录直接使用ftp.login()
        # ftp.cwd('县域视频数据统计通报'.encode('gbk').decode('iso-8859-1'))
        ftp.cwd(cwd.encode('gbk').decode('iso-8859-1'))
        fp = open(src, 'rb')
        ftp.storbinary('STOR ' + dst.encode('gbk').decode('iso-8859-1'), fp, 1024)
        print(str(dst) + '上传ftp ok')
    except Exception as e:
        print('上传ftp异常：' + str(e))
    finally:
        ftp.close()  # 关闭ftp

def get_zip(dir, zip_name):
    zip = zipfile.ZipFile(zip_name, 'w', zipfile.ZIP_DEFLATED)
    for root, dirs, files in os.walk(dir):
        for file in files:
            zip.write(os.path.join(root, file),file)
    zip.close()
    print(str(zip_name) + '压缩 ok')

mydb = pymysql.connect(
    host='13.32.4.170',
    # host='192.168.23.112',
    port=3306,
    user='root',
    password='wanfang@2001',
    db='db1400',
    # db='ligh',
    charset='utf8'
)


def mysql_select(mydb, sql):
    res_select = ()
    try:
        # print('查询...')
        cursor = mydb.cursor()
        cursor.execute(sql)
        res_select = cursor.fetchall()
    except Exception as e:
        print("查询异常:" + str(e))
    finally:
        cursor.close()
        return res_select


def open_xlsx(file):
    workbook = None
    try:
        # 打开excle文件，获取工作簿对象
        workbook = openpyxl.load_workbook(file)
    except Exception as e:
        print(file + '打开异常：' + str(e))
    finally:
        return workbook


def copy_file(src_file, dst_file):
    if not os.path.exists(dst_file):
        shutil.copy(src_file, dst_file)
        print(str(dst_file) + "创建成功！")


if __name__ == '__main__':
    weekday_int = datetime.datetime.now().isoweekday()
    border = Border(left=Side(border_style='thin', color='000000'), right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'), bottom=Side(border_style='thin', color='000000'))
    ftp_cwd = '智慧社区无效统计'
    today = time.strftime("%Y-%m-%d", time.localtime())
    home_dir = '/home/ligh/relay_stat/'
    # home_dir = 'C:/Users/wf/Desktop/'
    today_dir = os.path.join(home_dir, str(today))
    if not os.path.exists(today_dir):
        os.mkdir(today_dir)
    area_list = ['鹿泉区', '井陉县', '藁城区', '平山县', '行唐县', '无极县', '高邑县', '栾城区', '矿区', '长安区', '循环化工园区', '晋州市', '赵县', '高新区', '裕华区', '新华区', '深泽县', '正定县', '灵寿县', '新乐市', '元氏县', '桥西区', '赞皇县']

     #24区县无效设备统计
    dict_all = {}
    sql_invalid_ape="select b.xian, b.xqmc,a.name,a.ape_id  from ape a  join dm_xq b  on a.place_code = b.bm where not exists (select 1 from (select c.device_id ape_id  from t_push_data_log c  where c.type = '12' AND date=CURDATE() group by device_id) c  where a.ape_id = c.ape_id)  and a.function_type = '2' order by b.xian"
    res_invalid_ape = mysql_select(mydb, sql_invalid_ape)
    if res_invalid_ape:
        for res in res_invalid_ape:
            if res[0] not in dict_all.keys():
                dict_all.update({res[0]:[[],[]]})
                dict_all[res[0]][0].append(res[1::])
            else:
                dict_all[res[0]][0].append(res[1::])

    # 24区县无效卡口统计
    sql_invalid_toll = "select 	b.xian,	b.xqmc,	a. NAME,	a.tollgate_id FROM	tollgate a JOIN dm_xq b ON a.place_code = b.bm where NOT EXISTS (select 1 from (select c.tollgate_id tollgate_id from t_push_data_log c where c.type = '13'AND date = CURDATE() GROUP by tollgate_id ) c where a.tollgate_id = c.tollgate_id ) ORDER by b.xian"
    res_invalid_toll = mysql_select(mydb, sql_invalid_toll)
    if res_invalid_toll:
        for res in res_invalid_toll:
            if res[0] not in dict_all.keys():
                dict_all.update({res[0]: [[], []]})
                dict_all[res[0]][1].append(res[1::])
            else:
                dict_all[res[0]][1].append(res[1::])

    dict_summary = {}
    for k,v in dict_all.items():
        #汇总表
        dict_summary.update({k:[len(v[0]), len(v[1])]})

        workbook_invalid = Workbook()  # 创建工作簿
        # sheet页初始化
        sheet_ape = workbook_invalid.create_sheet('无效人脸设备')
        sheet_ape['A1'] = '序号'
        sheet_ape['B1'] = '区县'
        sheet_ape['C1'] = '小区名称'
        sheet_ape['D1'] = '设备名称'
        sheet_ape['E1'] = '设备ID'
        #设定边框
        sheet_ape['A1'].border = border
        sheet_ape['B1'].border = border
        sheet_ape['C1'].border = border
        sheet_ape['D1'].border = border
        sheet_ape['E1'].border = border

        sheet_toll = workbook_invalid.create_sheet('无效车辆卡口')
        sheet_toll['A1'] = '序号'
        sheet_toll['B1'] = '区县'
        sheet_toll['C1'] = '小区名称'
        sheet_toll['D1'] = '卡口名称'
        sheet_toll['E1'] = '卡口ID'
        # 设定边框
        sheet_toll['A1'].border = border
        sheet_toll['B1'].border = border
        sheet_toll['C1'].border = border
        sheet_toll['D1'].border = border
        sheet_toll['E1'].border = border

        # '无效设备'sheet页写入
        ape_num = 1
        for ape in v[0]:
            sheet_ape['A' + str(ape_num + 1)] = int(ape_num)
            sheet_ape['B' + str(ape_num + 1)] = str(k)
            sheet_ape['C' + str(ape_num + 1)] = str(ape[0])
            sheet_ape['D' + str(ape_num + 1)] = str(ape[1])
            sheet_ape['E' + str(ape_num + 1)] = str(ape[2])

            sheet_ape['A' + str(ape_num + 1)].border = border
            sheet_ape['B' + str(ape_num + 1)].border = border
            sheet_ape['C' + str(ape_num + 1)].border = border
            sheet_ape['D' + str(ape_num + 1)].border = border
            sheet_ape['E' + str(ape_num + 1)].border = border
            ape_num += 1

        # '无效设备'sheet页写入
        toll_num = 1
        for toll in v[1]:
            sheet_toll['A' + str(toll_num + 1)] = int(toll_num)
            sheet_toll['B' + str(toll_num + 1)] = str(k)
            sheet_toll['C' + str(toll_num + 1)] = str(toll[0])
            sheet_toll['D' + str(toll_num + 1)] = str(toll[1])
            sheet_toll['E' + str(toll_num + 1)] = str(toll[2])

            sheet_toll['A' + str(toll_num + 1)].border = border
            sheet_toll['B' + str(toll_num + 1)].border = border
            sheet_toll['C' + str(toll_num + 1)].border = border
            sheet_toll['D' + str(toll_num + 1)].border = border
            sheet_toll['E' + str(toll_num + 1)].border = border
            toll_num += 1

        # 保存、关闭excel文件
        workbook_invalid.remove(workbook_invalid['Sheet'])
        workbook_invalid.save(os.path.join(today_dir, k + '.xlsx'))
        workbook_invalid.close()
        print(k + ' ok')


    #23区县无效汇总.xlsx
    workbook_summary = Workbook()  # 创建工作簿
    # sheet页初始化
    sheet_summary = workbook_summary.create_sheet('无效汇总')
    sheet_summary['A1'] = '序号'
    sheet_summary['B1'] = '区县'
    sheet_summary['C1'] = '无效人脸设备数'
    sheet_summary['D1'] = '无效车辆卡口数'

    # 设定边框
    sheet_summary['A1'].border = border
    sheet_summary['B1'].border = border
    sheet_summary['C1'].border = border
    sheet_summary['D1'].border = border

    summary_num = 1
    for k,v in dict_summary.items():

        sheet_summary['A' + str(summary_num + 1)] = int(summary_num)
        sheet_summary['B' + str(summary_num + 1)] = str(k)
        sheet_summary['C' + str(summary_num + 1)] = v[0]
        sheet_summary['D' + str(summary_num + 1)] = v[1]

        sheet_summary['A' + str(summary_num + 1)].border = border
        sheet_summary['B' + str(summary_num + 1)].border = border
        sheet_summary['C' + str(summary_num + 1)].border = border
        sheet_summary['D' + str(summary_num + 1)].border = border
        summary_num += 1

    workbook_summary.remove(workbook_summary['Sheet'])
    workbook_summary.save(os.path.join(today_dir, '23区县无效汇总.xlsx'))
    workbook_summary.close()
    print('23区县无效汇总 ok')

    #压缩日期文件夹
    try:
        get_zip(today_dir, str(today) + '.zip')
    except Exception as e:
        print('zip压缩异常：', e)
    ftp_upload('智慧社区无效统计', os.path.join(home_dir, str(today) + '.zip'), str(today) + '.zip')
    # with open(os.path.join(home_dir, ".last.json"), "w") as fw:
    #     json.dump(dict_summary, fw)





    if weekday_int in [1, 4]:
        with open(os.path.join(home_dir, ".last.json"), 'r') as fr:
            dict_last = json.load(fr)
        for k, v in dict_summary.items():
            if k not in dict_last.keys():
                dict_summary.update({k: [0, 0]})
            dict_last.update({k:[v[0], v[0] - dict_last[k][0], v[1], v[1] - dict_last[k][1]]})
        # print(dict_last)

        # 环比：23区县无效汇总.xlsx
        workbook_relative = Workbook()  # 创建工作簿
        # sheet页初始化
        sheet_relative = workbook_relative.create_sheet('无效汇总环比')
        sheet_relative['A1'] = '序号'
        sheet_relative['B1'] = '区县'
        sheet_relative['C1'] = '无效人脸设备数'
        sheet_relative['D1'] = '人脸环比'
        sheet_relative['E1'] = '无效车辆卡口数'
        sheet_relative['F1'] = '车辆环比'

        # 设定边框
        sheet_relative['A1'].border = border
        sheet_relative['B1'].border = border
        sheet_relative['C1'].border = border
        sheet_relative['D1'].border = border
        sheet_relative['E1'].border = border
        sheet_relative['F1'].border = border

        relative_num = 1
        for k, v in dict_last.items():
            sheet_relative['A' + str(relative_num + 1)] = int(relative_num)
            sheet_relative['B' + str(relative_num + 1)] = str(k)
            sheet_relative['C' + str(relative_num + 1)] = v[0]
            sheet_relative['D' + str(relative_num + 1)] = v[1]
            sheet_relative['E' + str(relative_num + 1)] = v[2]
            sheet_relative['F' + str(relative_num + 1)] = v[3]

            sheet_relative['A' + str(relative_num + 1)].border = border
            sheet_relative['B' + str(relative_num + 1)].border = border
            sheet_relative['C' + str(relative_num + 1)].border = border
            sheet_relative['D' + str(relative_num + 1)].border = border
            sheet_relative['E' + str(relative_num + 1)].border = border
            sheet_relative['F' + str(relative_num + 1)].border = border
            relative_num += 1

        with open(os.path.join(home_dir, ".last.json"), "w") as fw:
            json.dump(dict_summary, fw)

        workbook_relative.remove(workbook_relative['Sheet'])
        workbook_relative.save(os.path.join(home_dir, '23区县无效汇总环比_' + str(today) + '.xlsx'))
        workbook_relative.close()
        print('23区县无效汇总环比汇总 ok')
        ftp_upload('智慧社区无效统计', os.path.join(home_dir, '23区县无效汇总环比_' + str(today) + '.xlsx'), '23区县无效汇总环比_' + str(today) + '.xlsx')


mydb.close()  # 关闭mysql连接
print('脚本结束!')